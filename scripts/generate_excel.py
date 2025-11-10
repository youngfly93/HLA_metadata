#!/usr/bin/env python3
"""
Excel报告生成脚本
生成包含多个工作表的最终元数据报告
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel formatting will be limited.")

# 项目路径配置
PROJECT_ROOT = Path(__file__).parent.parent
DATA_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
DATA_VALIDATION_DIR = PROJECT_ROOT / "data" / "validation"


class ExcelReportGenerator:
    """Excel报告生成器"""

    def __init__(self):
        self.output_file = DATA_PROCESSED_DIR / "proteomics_metadata_complete.xlsx"
        DATA_VALIDATION_DIR.mkdir(parents=True, exist_ok=True)

    def load_data(self) -> pd.DataFrame:
        """加载分类后的元数据（优先使用最新处理的数据）"""
        # 优先级：手动补充 > 交叉验证 > 推断后 > 清理后 > 分类后
        enriched_file = DATA_PROCESSED_DIR / "all_metadata_manually_enriched.csv"
        crosschecked_file = DATA_PROCESSED_DIR / "all_metadata_crosschecked.csv"
        inferred_file = DATA_PROCESSED_DIR / "all_metadata_inferred.csv"
        cleaned_file = DATA_PROCESSED_DIR / "all_metadata_cleaned.csv"
        classified_file = DATA_PROCESSED_DIR / "all_metadata_classified.csv"

        if enriched_file.exists():
            input_file = enriched_file
            print(f"Loading manually enriched data from: {input_file}")
        elif crosschecked_file.exists():
            input_file = crosschecked_file
            print(f"Loading cross-checked data from: {input_file}")
        elif inferred_file.exists():
            input_file = inferred_file
            print(f"Loading inferred data from: {input_file}")
        elif cleaned_file.exists():
            input_file = cleaned_file
            print(f"Loading cleaned data from: {input_file}")
        elif classified_file.exists():
            input_file = classified_file
            print(f"Loading classified data from: {input_file}")
        else:
            print(f"✗ Error: No metadata file found!")
            print("  Please run classify_metadata.py first")
            sys.exit(1)

        df = pd.read_csv(input_file)
        print(f"✓ Loaded {len(df)} datasets\n")

        return df

    def create_main_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        创建主元数据表

        包含所有数据集的关键信息
        """
        print("Creating main metadata sheet...")

        # 选择关键列
        main_columns = [
            'dataset_id',
            'repository',
            'title',
            'hla_type',
            'disease_type',
            'disease_category',
            'is_healthy',
            'sample_type',
            'organisms',
            'tissues',
            'cell_types',
            'instruments',
            'ptms',
            'project_tags',
            'keywords',
            'publication_date',
            'pubmed_ids',
            'pride_url',
            'metadata_quality',
            'needs_manual_review',
        ]

        # 只保留存在的列
        available_columns = [col for col in main_columns if col in df.columns]
        main_df = df[available_columns].copy()

        # 重命名列为中英文对照
        column_names = {
            'dataset_id': '数据集ID',
            'repository': '数据库',
            'title': '标题',
            'hla_type': 'HLA类型',
            'disease_type': '疾病类型',
            'disease_category': '疾病类别',
            'is_healthy': '健康对照',
            'sample_type': '样本类型',
            'organisms': '物种',
            'tissues': '组织',
            'cell_types': '细胞类型',
            'instruments': '质谱仪器',
            'ptms': '翻译后修饰',
            'project_tags': '项目标签',
            'keywords': '关键词',
            'publication_date': '发表日期',
            'pubmed_ids': 'PubMed ID',
            'pride_url': 'URL',
            'metadata_quality': '元数据质量',
            'needs_manual_review': '需人工审核',
        }

        main_df = main_df.rename(columns=column_names)

        print(f"  ✓ Main sheet created with {len(main_df)} rows")
        return main_df

    def create_disease_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建疾病类型汇总表"""
        print("Creating disease summary sheet...")

        # 疾病类别统计
        disease_category_counts = df['disease_category'].value_counts().reset_index()
        disease_category_counts.columns = ['疾病类别', '数据集数量']
        disease_category_counts['百分比'] = (
            disease_category_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # 具体疾病类型统计（前30个）
        disease_type_counts = df['disease_type'].value_counts().head(30).reset_index()
        disease_type_counts.columns = ['疾病类型', '数据集数量']
        disease_type_counts['百分比'] = (
            disease_type_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # 健康对照统计
        healthy_counts = df['is_healthy'].value_counts().reset_index()
        healthy_counts.columns = ['是否健康对照', '数据集数量']
        healthy_counts['是否健康对照'] = healthy_counts['是否健康对照'].map(
            {True: '是', False: '否'}
        )

        # 合并表格（添加空行分隔）
        summary_df = pd.DataFrame()
        summary_df = pd.concat([
            pd.DataFrame([['疾病类别统计', '', '']]),
            disease_category_counts,
            pd.DataFrame([['', '', '']]),
            pd.DataFrame([['健康对照统计', '', '']]),
            healthy_counts,
            pd.DataFrame([['', '', '']]),
            pd.DataFrame([['具体疾病类型统计（前30）', '', '']]),
            disease_type_counts,
        ], ignore_index=True)

        print(f"  ✓ Disease summary created")
        return summary_df

    def create_hla_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建HLA分类汇总表"""
        print("Creating HLA summary sheet...")

        # HLA类型统计
        hla_counts = df['hla_type'].value_counts().reset_index()
        hla_counts.columns = ['HLA类型', '数据集数量']
        hla_counts['百分比'] = (
            hla_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # 需要人工审核的HLA数据集
        needs_review = df[df['hla_needs_review'] == True][
            ['dataset_id', 'title', 'hla_type', 'project_tags', 'keywords']
        ].copy()
        needs_review.columns = [
            '数据集ID', '标题', 'HLA类型', '项目标签', '关键词'
        ]

        # 合并表格
        summary_df = pd.concat([
            pd.DataFrame([['HLA类型统计', '', '']]),
            hla_counts,
            pd.DataFrame([['', '', '']]),
            pd.DataFrame([['需人工确认的HLA数据集', '', '', '', '']]),
            needs_review,
        ], ignore_index=True)

        print(f"  ✓ HLA summary created")
        return summary_df

    def create_sample_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建样本类型汇总表"""
        print("Creating sample type summary sheet...")

        # 样本类型统计
        sample_counts = df['sample_type'].value_counts().reset_index()
        sample_counts.columns = ['样本类型', '数据集数量']
        sample_counts['百分比'] = (
            sample_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # 按类别分组
        blood_samples = df[df['sample_type'].str.contains('Blood', na=False)][
            ['dataset_id', 'sample_type', 'tissues', 'title']
        ].copy()
        blood_samples.columns = ['数据集ID', '样本类型', '组织详情', '标题']

        tissue_samples = df[df['sample_type'].str.contains('Tissue', na=False)][
            ['dataset_id', 'sample_type', 'tissues', 'title']
        ].head(50).copy()
        tissue_samples.columns = ['数据集ID', '样本类型', '组织详情', '标题']

        cell_line_samples = df[df['sample_type'].str.contains('Cell line', na=False)][
            ['dataset_id', 'sample_type', 'cell_types', 'title']
        ].copy()
        cell_line_samples.columns = ['数据集ID', '样本类型', '细胞详情', '标题']

        # 合并表格
        summary_df = pd.concat([
            pd.DataFrame([['样本类型统计', '', '']]),
            sample_counts,
            pd.DataFrame([['', '', '', '']]),
            pd.DataFrame([['血液样本详情', '', '', '']]),
            blood_samples,
            pd.DataFrame([['', '', '', '']]),
            pd.DataFrame([['细胞系样本详情', '', '', '']]),
            cell_line_samples,
            pd.DataFrame([['', '', '', '']]),
            pd.DataFrame([['组织样本详情（前50）', '', '', '']]),
            tissue_samples,
        ], ignore_index=True)

        print(f"  ✓ Sample type summary created")
        return summary_df

    def create_quality_report(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建数据质量报告"""
        print("Creating quality report sheet...")

        # 数据质量统计
        quality_counts = df['metadata_quality'].value_counts().reset_index()
        quality_counts.columns = ['质量等级', '数据集数量']
        quality_counts['百分比'] = (
            quality_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # 数据库分布
        repo_counts = df['repository'].value_counts().reset_index()
        repo_counts.columns = ['数据库', '数据集数量']
        repo_counts['百分比'] = (
            repo_counts['数据集数量'] / len(df) * 100
        ).round(2)

        # SDRF文件统计
        if 'has_sdrf' in df.columns:
            sdrf_counts = df['has_sdrf'].value_counts().reset_index()
            sdrf_counts.columns = ['有SDRF文件', '数据集数量']
            sdrf_counts['有SDRF文件'] = sdrf_counts['有SDRF文件'].map(
                {True: '是', False: '否'}
            )
        else:
            sdrf_counts = pd.DataFrame([['统计不可用', 0]], columns=['有SDRF文件', '数据集数量'])

        # 需要人工审核的数据集列表
        manual_review = df[df['needs_manual_review'] == True][
            ['dataset_id', 'repository', 'hla_type', 'sample_type',
             'disease_type', 'metadata_quality']
        ].copy()
        manual_review.columns = [
            '数据集ID', '数据库', 'HLA类型', '样本类型', '疾病类型', '质量等级'
        ]

        # 合并表格
        summary_df = pd.concat([
            pd.DataFrame([['数据质量统计', '', '']]),
            quality_counts,
            pd.DataFrame([['', '', '']]),
            pd.DataFrame([['数据库分布', '', '']]),
            repo_counts,
            pd.DataFrame([['', '', '']]),
            pd.DataFrame([['SDRF文件统计', '', '']]),
            sdrf_counts,
            pd.DataFrame([['', '', '', '', '', '']]),
            pd.DataFrame([['需人工审核的数据集', '', '', '', '', '']]),
            manual_review,
        ], ignore_index=True)

        print(f"  ✓ Quality report created")
        return summary_df

    def create_technical_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建技术信息汇总表"""
        print("Creating technical summary sheet...")

        # 质谱仪器统计
        instruments_data = []
        for instruments_str in df['instruments'].dropna():
            if instruments_str and str(instruments_str) != 'nan':
                for instrument in str(instruments_str).split(';'):
                    instruments_data.append(instrument.strip())

        if instruments_data:
            instrument_counts = pd.Series(instruments_data).value_counts().head(20).reset_index()
            instrument_counts.columns = ['质谱仪器', '使用次数']
        else:
            instrument_counts = pd.DataFrame([['无数据', 0]], columns=['质谱仪器', '使用次数'])

        # PTM统计
        ptms_data = []
        for ptms_str in df['ptms'].dropna():
            if ptms_str and str(ptms_str) != 'nan':
                for ptm in str(ptms_str).split(';'):
                    ptms_data.append(ptm.strip())

        if ptms_data:
            ptm_counts = pd.Series(ptms_data).value_counts().head(20).reset_index()
            ptm_counts.columns = ['翻译后修饰类型', '出现次数']
        else:
            ptm_counts = pd.DataFrame([['无数据', 0]], columns=['翻译后修饰类型', '出现次数'])

        # 物种统计
        organisms_data = []
        for org_str in df['organisms'].dropna():
            if org_str and str(org_str) != 'nan':
                for org in str(org_str).split(';'):
                    organisms_data.append(org.strip())

        if organisms_data:
            organism_counts = pd.Series(organisms_data).value_counts().head(15).reset_index()
            organism_counts.columns = ['物种', '数据集数量']
        else:
            organism_counts = pd.DataFrame([['无数据', 0]], columns=['物种', '数据集数量'])

        # 合并表格
        summary_df = pd.concat([
            pd.DataFrame([['质谱仪器统计（前20）', '']]),
            instrument_counts,
            pd.DataFrame([['', '']]),
            pd.DataFrame([['翻译后修饰统计（前20）', '']]),
            ptm_counts,
            pd.DataFrame([['', '']]),
            pd.DataFrame([['物种分布（前15）', '']]),
            organism_counts,
        ], ignore_index=True)

        print(f"  ✓ Technical summary created")
        return summary_df

    def format_excel(self, writer):
        """格式化Excel文件（如果openpyxl可用）"""
        if not OPENPYXL_AVAILABLE:
            return

        print("\nApplying Excel formatting...")

        workbook = writer.book

        # 定义样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)

        # 格式化每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 设置标题行样式
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 自动调整列宽
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 冻结首行
            sheet.freeze_panes = sheet['A2']

        print("  ✓ Excel formatting applied")

    def generate_report(self):
        """生成完整的Excel报告"""
        print("\n" + "="*60)
        print("Generating Excel Report")
        print("="*60 + "\n")

        # 加载数据
        df = self.load_data()

        # 创建Excel writer
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Sheet 1: 主元数据表
            main_df = self.create_main_sheet(df)
            main_df.to_excel(writer, sheet_name='主元数据表', index=False)

            # Sheet 2: 疾病类型汇总
            disease_df = self.create_disease_summary(df)
            disease_df.to_excel(writer, sheet_name='疾病类型汇总', index=False, header=False)

            # Sheet 3: HLA分类汇总
            hla_df = self.create_hla_summary(df)
            hla_df.to_excel(writer, sheet_name='HLA分类汇总', index=False, header=False)

            # Sheet 4: 样本类型汇总
            sample_df = self.create_sample_summary(df)
            sample_df.to_excel(writer, sheet_name='样本类型汇总', index=False, header=False)

            # Sheet 5: 技术信息汇总
            tech_df = self.create_technical_summary(df)
            tech_df.to_excel(writer, sheet_name='技术信息汇总', index=False, header=False)

            # Sheet 6: 数据质量报告
            quality_df = self.create_quality_report(df)
            quality_df.to_excel(writer, sheet_name='数据质量报告', index=False, header=False)

            # 应用格式
            self.format_excel(writer)

        print("\n" + "="*60)
        print("Report Generation Summary")
        print("="*60)
        print(f"Total datasets: {len(df)}")
        print(f"Output file: {self.output_file}")
        print(f"File size: {self.output_file.stat().st_size / 1024:.2f} KB")
        print("\n✓ Excel report generated successfully!")
        print(f"\nYou can now open: {self.output_file}")
        print("\n")

        # 生成质量报告文本文件
        self.generate_validation_report(df)

    def generate_validation_report(self, df: pd.DataFrame):
        """生成验证报告文本文件"""
        report_file = DATA_VALIDATION_DIR / "quality_report.txt"

        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("="*60 + "\n")
            f.write("HLA元数据收集质量报告\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n\n")

            f.write(f"总数据集数量: {len(df)}\n\n")

            f.write("数据库分布:\n")
            for repo, count in df['repository'].value_counts().items():
                f.write(f"  {repo}: {count}\n")
            f.write("\n")

            f.write("HLA类型分布:\n")
            for hla_type, count in df['hla_type'].value_counts().items():
                f.write(f"  {hla_type}: {count}\n")
            f.write("\n")

            f.write("样本类型分布:\n")
            for sample_type, count in df['sample_type'].value_counts().head(10).items():
                f.write(f"  {sample_type}: {count}\n")
            f.write("\n")

            f.write("疾病类别分布:\n")
            for disease_cat, count in df['disease_category'].value_counts().items():
                f.write(f"  {disease_cat}: {count}\n")
            f.write("\n")

            f.write("数据质量分布:\n")
            for quality, count in df['metadata_quality'].value_counts().items():
                f.write(f"  {quality}: {count}\n")
            f.write("\n")

            review_count = df['needs_manual_review'].sum()
            f.write(f"需要人工审核的数据集: {review_count}\n")

            if review_count > 0:
                f.write("\n需人工审核的数据集列表:\n")
                for idx, row in df[df['needs_manual_review'] == True].iterrows():
                    f.write(f"  - {row['dataset_id']}: {row['hla_type']}\n")

        print(f"✓ Validation report saved to: {report_file}")


def main():
    """主函数"""
    generator = ExcelReportGenerator()
    generator.generate_report()


if __name__ == "__main__":
    main()
